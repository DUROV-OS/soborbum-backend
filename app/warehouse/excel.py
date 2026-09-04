import io

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook

from app.warehouse.models import MaterialCategory

TEMPLATE_HEADERS = [
    "ID материала (если уже есть на складе)",
    "Категория",
    "Код",
    "Название",
    "Единица измерения",
    "Может быть дробной (да/нет)",
    "Цена закупки",
    "Количество поставки",
]

CATEGORY_BY_LABEL = {category.value: category for category in MaterialCategory}


def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Поставка"
    ws.append(TEMPLATE_HEADERS)
    ws.append([None, "брусы/доска", "BR-150x50-6000", "Брус 150х50х6000", "шт", "нет", 850, 100])
    ws.append([1, None, None, None, None, None, None, 50])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_bool(value, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("да", "yes", "true", "1")


def parse_supply_rows(file: UploadFile) -> list[dict]:
    try:
        content = file.file.read()
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не удалось прочитать excel-файл")

    ws = wb.active
    rows: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=8, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        material_id, category_label, code, title, unit, is_fractional, purchase_price, quantity = (
            list(row) + [None] * 8
        )[:8]
        if quantity is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Строка {idx}: не указано количество")

        category = MaterialCategory.NONE
        if category_label not in (None, ""):
            category = CATEGORY_BY_LABEL.get(str(category_label).strip())
            if category is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Строка {idx}: неизвестная категория «{category_label}»",
                )

        if material_id in (None, ""):
            if not title or not code:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Строка {idx}: для нового материала нужно указать код и название",
                )

        rows.append(
            {
                "warehouse_material_id": int(material_id) if material_id not in (None, "") else None,
                "category": category,
                "code": code,
                "title": title,
                "unit": unit,
                "is_fractional": _parse_bool(is_fractional),
                "purchase_price": float(purchase_price) if purchase_price not in (None, "") else 0,
                "quantity": float(quantity),
            }
        )
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="В файле нет ни одной строки с данными")
    return rows
